#!/usr/bin/env python3
"""
BTC International Participation Tracker — Data Builder
======================================================

Reads ``data/tracker.xlsx`` (one sheet per season) and writes
``data/seasons.json``, the structured payload consumed by the dashboard.

Usage:
    python scripts/build.py

The script is tolerant of the workbook's per-section layout
(Section 1 — Horses by Competition, Section 2 — Horse Register,
Section 3 — Winners, Section 4 — Trainers, Section 5 — Additional Races).
Trailing whitespace and inconsistent country labels are normalised.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "openpyxl is required.\n"
        "Install with:  pip install openpyxl\n"
    )
    sys.exit(1)


ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "tracker.xlsx"
DEST = ROOT / "data" / "seasons.json"


# ---------------------------------------------------------------------------
# Normalisers
# ---------------------------------------------------------------------------

COUNTRY_MAP = {
    "great britain": "GB", "england": "GB", "uk": "GB", "gb": "GB",
    "ireland": "IRE", "ire": "IRE",
    "france": "FR", "fr": "FR",
    "germany": "GER", "ger": "GER",
    "uae": "UAE", "united arab emirates": "UAE",
    "japan": "JPN", "jpn": "JPN",
    "norway": "NOR", "nor": "NOR",
    "qatar": "QAT", "qat": "QAT", "qtr": "QAT",
    "usa": "USA", "united states": "USA",
    "bahrain": "BHR", "bhr": "BHR",
    "denmark": "DEN", "den": "DEN",
}

COMPETITION_MAP = {
    # canonical short form -> longer label is generated in JS if needed
    "bahrain international trophy": "BIT",
    "bit": "BIT",
    "bahrain turf series": "BTS",
    "baharain turf series": "BTS",   # typo seen in source
    "bahrain turs series": "BTS",    # typo seen in source
    "bts": "BTS",
    "hh shaikh nasser cup": "HH Shaikh Nasser Cup",
    "hh shaikh nasser day (bts)": "HH Shaikh Nasser Cup",
    "crown prince cup": "Crown Prince Cup",
    "cp cup": "Crown Prince Cup",
    "cpc": "Crown Prince Cup",
    "king's cup": "King's Cup",
    "kings cup": "King's Cup",
    "kc": "King's Cup",
    "additional": "Additional",
}


def clean(value: Any) -> str:
    """Strip whitespace and collapse internal multi-spaces."""
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_country(value: Any) -> str:
    s = clean(value)
    if not s:
        return ""
    return COUNTRY_MAP.get(s.lower(), s.upper() if len(s) <= 4 else s)


def norm_competition(value: Any) -> str:
    s = clean(value)
    if not s:
        return ""
    key = s.lower()
    # Composite labels like "Baharain Turf Series + HH Shaikh Nasser Cup"
    if "+" in key:
        parts = [norm_competition(p) for p in s.split("+")]
        return " + ".join(p for p in parts if p)
    if key in COMPETITION_MAP:
        return COMPETITION_MAP[key]
    # Heuristics for partial matches
    if "international trophy" in key or key.startswith("bit"):
        return "BIT"
    if "turf series" in key or "turs series" in key:
        return "BTS"
    if "shaikh nasser" in key:
        return "HH Shaikh Nasser Cup"
    if "crown prince" in key:
        return "Crown Prince Cup"
    if "king" in key and "cup" in key:
        return "King's Cup"
    return s


def split_horse_origin(raw: str) -> tuple[str, str]:
    """``Royal Champion (IRE)`` -> ('Royal Champion', 'IRE')."""
    s = clean(raw)
    m = re.match(r"^(.*?)\s*\(([A-Za-z]{2,4})\)\s*$", s)
    if m:
        return m.group(1).strip(), m.group(2).upper()
    return s, ""


def to_iso_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = clean(value)
    # Try DD/MM/YY or DD/MM/YYYY
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b", "%d-%b-%y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year < 100:
                dt = dt.replace(year=dt.year + 2000)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # leave as-is if unparseable


# ---------------------------------------------------------------------------
# Section parsers
# ---------------------------------------------------------------------------

SECTION_HEADERS = {
    "section 1": "by_competition",
    "section 2": "register",
    "section 3": "winners",
    "section 4": "trainers",
    "section 5": "additional",
}


def get_rows(ws) -> list[list[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def find_sections(rows: list[list[Any]]) -> dict[str, tuple[int, int]]:
    """Return {key: (start_row, end_row_exclusive)} for the five sections."""
    indices: list[tuple[str, int]] = []
    for i, row in enumerate(rows):
        if not row:
            continue
        first = clean(row[0]).lower()
        for prefix, key in SECTION_HEADERS.items():
            if first.startswith(prefix):
                indices.append((key, i))
                break
    sections: dict[str, tuple[int, int]] = {}
    for j, (key, start) in enumerate(indices):
        end = indices[j + 1][1] if j + 1 < len(indices) else len(rows)
        sections[key] = (start + 1, end)  # skip the section title row
    return sections


def parse_by_competition(rows: list[list[Any]], span: tuple[int, int]) -> list[dict]:
    start, end = span
    out: list[dict] = []
    seen_header = False
    for row in rows[start:end]:
        if not row or all(c is None for c in row):
            continue
        first = clean(row[0])
        if not seen_header:
            if first.lower() == "competition":
                seen_header = True
            continue
        if first.upper() == "TOTAL":
            continue
        comp = norm_competition(first)
        country = norm_country(row[1] if len(row) > 1 else "")
        count = row[2] if len(row) > 2 else None
        if comp and country and isinstance(count, (int, float)):
            out.append({"comp": comp, "country": country, "horses": int(count)})
    return out


def parse_register(rows: list[list[Any]], span: tuple[int, int]) -> list[dict]:
    start, end = span
    out: list[dict] = []
    seen_header = False
    for row in rows[start:end]:
        if not row or all(c is None for c in row):
            continue
        first = clean(row[0])
        if not first:
            continue
        if not seen_header:
            if first.lower() == "horse name":
                seen_header = True
            continue
        if first.upper().startswith("TOTAL"):
            continue
        horse_raw = first
        name, origin = split_horse_origin(horse_raw)
        trainer = clean(row[1]) if len(row) > 1 else ""
        country = norm_country(row[2] if len(row) > 2 else "")
        comp = norm_competition(row[3] if len(row) > 3 else "")
        winner_flag = clean(row[4]).upper() if len(row) > 4 else ""
        wins_val = row[5] if len(row) > 5 else None
        try:
            wins = int(wins_val) if wins_val not in (None, "") else 0
        except (TypeError, ValueError):
            wins = 0
        note = clean(row[6]) if len(row) > 6 else ""
        out.append({
            "name": name,
            "origin": origin,
            "trainer": trainer,
            "country": country,
            "comp": comp,
            "winner": winner_flag == "Y",
            "wins": wins,
            "note": note,
        })
    return out


def parse_winners(rows: list[list[Any]], span: tuple[int, int]) -> list[dict]:
    start, end = span
    out: list[dict] = []
    seen_header = False
    for row in rows[start:end]:
        if not row or all(c is None for c in row):
            continue
        first = clean(row[0])
        if not first:
            continue
        if not seen_header:
            if first.lower() == "horse":
                seen_header = True
            continue
        if first.lower().startswith(("total international", "individual winning", "dual")):
            continue
        name, origin = split_horse_origin(first)
        trainer = clean(row[1]) if len(row) > 1 else ""
        country = norm_country(row[2] if len(row) > 2 else "")
        race = clean(row[3]) if len(row) > 3 else ""
        date_iso = to_iso_date(row[4] if len(row) > 4 else "")
        win_no_raw = row[5] if len(row) > 5 else None
        # Some sheets used '1st' instead of 1
        if isinstance(win_no_raw, str):
            m = re.match(r"(\d+)", win_no_raw)
            win_no = int(m.group(1)) if m else 1
        else:
            try:
                win_no = int(win_no_raw) if win_no_raw not in (None, "") else 1
            except (TypeError, ValueError):
                win_no = 1
        note = clean(row[6]) if len(row) > 6 else ""
        out.append({
            "name": name,
            "origin": origin,
            "trainer": trainer,
            "country": country,
            "race": race,
            "date": date_iso,
            "win_no": win_no,
            "note": note,
        })
    return out


def parse_trainers(rows: list[list[Any]], span: tuple[int, int]) -> list[dict]:
    """Trainers section can have sub-headers like 'Bahrain International Trophy'."""
    start, end = span
    out: list[dict] = []
    current_comp = ""
    for row in rows[start:end]:
        if not row or all(c is None for c in row):
            continue
        first = clean(row[0])
        if not first:
            continue
        lower = first.lower()
        if lower.startswith("total international"):
            continue
        if lower == "trainer":  # header row
            continue
        # Sub-header lines
        if (lower.startswith(("bahrain international", "bahrain turf",
                              "hh shaikh", "crown prince", "king"))
                and (len(row) < 2 or row[1] in (None, ""))):
            current_comp = norm_competition(first)
            continue
        # Otherwise a trainer row
        name = first
        country = norm_country(row[1] if len(row) > 1 else "")
        note = clean(row[2]) if len(row) > 2 else ""
        out.append({
            "name": name,
            "country": country,
            "comp": current_comp,
            "note": note,
        })
    return out


def parse_additional(rows: list[list[Any]], span: tuple[int, int]) -> list[dict]:
    start, end = span
    out: list[dict] = []
    seen_header = False
    for row in rows[start:end]:
        if not row or all(c is None for c in row):
            continue
        first = clean(row[0])
        if not first:
            continue
        if not seen_header:
            if first.lower() == "race name":
                seen_header = True
            continue
        race = first
        horse_raw = clean(row[1]) if len(row) > 1 else ""
        name, origin = split_horse_origin(horse_raw)
        trainer = clean(row[2]) if len(row) > 2 else ""
        country = norm_country(row[3] if len(row) > 3 else "")
        date_iso = to_iso_date(row[4] if len(row) > 4 else "")
        result = clean(row[5]) if len(row) > 5 else ""
        note = clean(row[6]) if len(row) > 6 else ""
        out.append({
            "race": race,
            "name": name,
            "origin": origin,
            "trainer": trainer,
            "country": country,
            "date": date_iso,
            "result": result,
            "note": note,
        })
    return out


# ---------------------------------------------------------------------------
# Per-season aggregate stats (computed from the register so totals match UI)
# ---------------------------------------------------------------------------

def derive_stats(register: list[dict], winners: list[dict]) -> dict:
    horses = len(register)
    countries = sorted({h["country"] for h in register if h["country"]})
    trainers = sorted({h["trainer"] for h in register if h["trainer"]})
    win_horses = [h for h in register if h["winner"]]
    individual_winners = len(win_horses)
    total_wins = sum(h["wins"] for h in win_horses)
    dual_winners = sum(1 for h in win_horses if h["wins"] >= 2)
    return {
        "horses": horses,
        "wins": total_wins,
        "individual_winners": individual_winners,
        "dual_winners": dual_winners,
        "trainers": len(trainers),
        "countries": countries,
        "winners_logged": len(winners),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build() -> dict:
    if not SRC.exists():
        sys.stderr.write(f"Source XLSX not found: {SRC}\n")
        sys.exit(1)
    wb = load_workbook(SRC, data_only=True)

    season_sheets = [s for s in wb.sheetnames if re.match(r"^\d{4}-\d{2}$", s)]
    seasons_out: list[dict] = []

    for sheet_name in season_sheets:
        ws = wb[sheet_name]
        rows = get_rows(ws)
        sections = find_sections(rows)

        by_comp = parse_by_competition(rows, sections["by_competition"]) if "by_competition" in sections else []
        register = parse_register(rows, sections["register"]) if "register" in sections else []
        winners = parse_winners(rows, sections["winners"]) if "winners" in sections else []
        trainers = parse_trainers(rows, sections["trainers"]) if "trainers" in sections else []
        additional = parse_additional(rows, sections["additional"]) if "additional" in sections else []

        season_label = sheet_name.replace("-", "/20") if not sheet_name.startswith("19") else sheet_name
        # 2019-20 -> 2019/20, 2025-26 -> 2025/26
        season_label = f"{sheet_name[:4]}/{sheet_name[5:7]}"

        seasons_out.append({
            "season": season_label,
            "sheet": sheet_name,
            "by_competition": by_comp,
            "register": register,
            "winners": winners,
            "trainers": trainers,
            "additional": additional,
            "stats": derive_stats(register, winners),
        })

    # Cross-season aggregates
    total_horses = sum(s["stats"]["horses"] for s in seasons_out)
    total_wins = sum(s["stats"]["wins"] for s in seasons_out)
    total_individual_winners = sum(s["stats"]["individual_winners"] for s in seasons_out)
    all_countries = sorted({c for s in seasons_out for c in s["stats"]["countries"]})

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "data/tracker.xlsx",
        "seasons": seasons_out,
        "totals": {
            "seasons": len(seasons_out),
            "horses": total_horses,
            "wins": total_wins,
            "individual_winners": total_individual_winners,
            "countries": all_countries,
        },
    }
    return payload


def main() -> None:
    payload = build()
    DEST.parent.mkdir(parents=True, exist_ok=True)
    with open(DEST, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"✓ Wrote {DEST.relative_to(ROOT)}")
    print(f"  Seasons:           {len(payload['seasons'])}")
    print(f"  Total horses:      {payload['totals']['horses']}")
    print(f"  Total int'l wins:  {payload['totals']['wins']}")
    print(f"  Countries:         {', '.join(payload['totals']['countries'])}")
    for s in payload["seasons"]:
        st = s["stats"]
        print(f"   · {s['season']}: {st['horses']:>3} horses · "
              f"{st['wins']:>2} wins · {st['individual_winners']:>2} winners · "
              f"{len(st['countries'])} countries")


if __name__ == "__main__":
    main()
